"""FastAPI admin routes."""

import csv
import io
import os
import uuid
from datetime import datetime, timedelta
from typing import Any

from fastapi import APIRouter, Query, Request
from fastapi.responses import FileResponse, JSONResponse

from utils.database import get_db

router = APIRouter(prefix="/admin", tags=["admin"])

DATA_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "data"))


def _error(message: str, status_code: int) -> JSONResponse:
    return JSONResponse(status_code=status_code, content={"success": False, "message": message})


async def _read_json(request: Request) -> dict[str, Any]:
    try:
        data = await request.json()
    except Exception:
        return {}
    return data if isinstance(data, dict) else {}


def _is_admin(username: str) -> bool:
    if not username:
        return False
    conn = get_db()
    try:
        row = conn.execute("SELECT role FROM users WHERE username=?", (username,)).fetchone()
        return bool(row and row["role"] == "admin")
    finally:
        conn.close()


def _to_float(value: Any, default: float = 0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


@router.get("/devices")
async def list_devices():
    conn = get_db()
    try:
        rows = conn.execute("SELECT * FROM devices ORDER BY created_at DESC").fetchall()
    finally:
        conn.close()
    return {"success": True, "devices": [dict(row) for row in rows]}


@router.post("/devices")
async def add_device(request: Request):
    data = await _read_json(request)
    operator = str(data.get("operator", "")).strip()
    if not _is_admin(operator):
        return _error("Admin permission required.", 403)

    name = str(data.get("name", "")).strip()
    if not name:
        return _error("Device name is required.", 400)

    device_id = str(uuid.uuid4())[:8]
    device_type = str(data.get("type", "sensor")).strip() or "sensor"
    location = str(data.get("location", "")).strip()
    lat = _to_float(data.get("lat", 0), 0.0)
    lon = _to_float(data.get("lon", 0), 0.0)

    conn = get_db()
    try:
        conn.execute(
            """
            INSERT INTO devices (id, name, location, type, lat, lon, status, created_by)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (device_id, name, location, device_type, lat, lon, "online", operator),
        )
        conn.commit()
        row = conn.execute("SELECT * FROM devices WHERE id=?", (device_id,)).fetchone()
    finally:
        conn.close()

    return {"success": True, "message": "Device created.", "device": dict(row)}


@router.put("/devices/{device_id}")
async def update_device(device_id: str, request: Request):
    data = await _read_json(request)
    operator = str(data.get("operator", "")).strip()
    if not _is_admin(operator):
        return _error("Admin permission required.", 403)

    conn = get_db()
    try:
        existing = conn.execute("SELECT * FROM devices WHERE id=?", (device_id,)).fetchone()
        if not existing:
            return _error("Device not found.", 404)

        updates: dict[str, Any] = {}
        for key in ("name", "location", "type", "status"):
            if key in data and data.get(key) is not None:
                updates[key] = str(data.get(key))
        for key in ("lat", "lon"):
            if key in data and data.get(key) is not None:
                updates[key] = _to_float(data.get(key), 0.0)

        if updates:
            set_clause = ", ".join([f"{column}=?" for column in updates])
            values = list(updates.values()) + [device_id]
            conn.execute(f"UPDATE devices SET {set_clause} WHERE id=?", values)
            conn.commit()

        row = conn.execute("SELECT * FROM devices WHERE id=?", (device_id,)).fetchone()
    finally:
        conn.close()

    return {"success": True, "message": "Device updated.", "device": dict(row)}


@router.delete("/devices/{device_id}")
async def delete_device(device_id: str, operator: str = Query("", description="Operator username")):
    if not _is_admin(operator.strip()):
        return _error("Admin permission required.", 403)

    conn = get_db()
    try:
        row = conn.execute("SELECT id FROM devices WHERE id=?", (device_id,)).fetchone()
        if not row:
            return _error("Device not found.", 404)
        conn.execute("DELETE FROM devices WHERE id=?", (device_id,))
        conn.commit()
    finally:
        conn.close()

    return {"success": True, "message": "Device deleted."}


@router.get("/alarms")
async def list_alarms(
    status: str = Query("", description="Status filter"),
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
):
    offset = (page - 1) * per_page

    conn = get_db()
    try:
        if status and status != "all":
            total = conn.execute("SELECT COUNT(*) AS c FROM alarms WHERE status=?", (status,)).fetchone()["c"]
            rows = conn.execute(
                "SELECT * FROM alarms WHERE status=? ORDER BY time DESC LIMIT ? OFFSET ?",
                (status, per_page, offset),
            ).fetchall()
        else:
            total = conn.execute("SELECT COUNT(*) AS c FROM alarms").fetchone()["c"]
            rows = conn.execute(
                "SELECT * FROM alarms ORDER BY time DESC LIMIT ? OFFSET ?",
                (per_page, offset),
            ).fetchall()
    finally:
        conn.close()

    return {
        "success": True,
        "alarms": [dict(row) for row in rows],
        "total": total,
        "page": page,
        "per_page": per_page,
    }


@router.post("/alarms")
async def add_alarm(request: Request):
    data = await _read_json(request)

    alarm_id = str(uuid.uuid4())[:8]
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    station = str(data.get("station", "unknown"))
    value = _to_float(data.get("value", 0), 0.0)
    threshold = _to_float(data.get("threshold", 0), 0.0)
    unit = str(data.get("unit", "ppb"))
    alarm_time = str(data.get("time", now))

    conn = get_db()
    try:
        conn.execute(
            """
            INSERT INTO alarms (id, station, value, threshold, unit, time, status)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (alarm_id, station, value, threshold, unit, alarm_time, "pending"),
        )
        conn.commit()
        row = conn.execute("SELECT * FROM alarms WHERE id=?", (alarm_id,)).fetchone()
    finally:
        conn.close()

    return {"success": True, "alarm": dict(row)}


@router.put("/alarms/{alarm_id}")
async def handle_alarm(alarm_id: str, request: Request):
    data = await _read_json(request)
    handler = str(data.get("handler", "")).strip()
    remark = str(data.get("remark", "")).strip()

    conn = get_db()
    try:
        row = conn.execute("SELECT * FROM alarms WHERE id=?", (alarm_id,)).fetchone()
        if not row:
            return _error("Alarm not found.", 404)

        conn.execute(
            """
            UPDATE alarms
            SET status='resolved', handler=?, handle_time=?, remark=?
            WHERE id=?
            """,
            (handler, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), remark, alarm_id),
        )
        conn.commit()
        updated = conn.execute("SELECT * FROM alarms WHERE id=?", (alarm_id,)).fetchone()
    finally:
        conn.close()

    return {"success": True, "message": "Alarm resolved.", "alarm": dict(updated)}


@router.get("/history")
async def list_history(
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    operator: str = Query("", description="Operator filter"),
    search: str = Query("", description="Search keyword"),
):
    offset = (page - 1) * per_page
    where_parts: list[str] = []
    params: list[Any] = []

    operator = operator.strip()
    search = search.strip()
    if operator:
        where_parts.append("operator=?")
        params.append(operator)
    if search:
        where_parts.append("(gas_type LIKE ? OR remark LIKE ?)")
        params.extend([f"%{search}%", f"%{search}%"])

    where_clause = ""
    if where_parts:
        where_clause = " WHERE " + " AND ".join(where_parts)

    conn = get_db()
    try:
        total = conn.execute(f"SELECT COUNT(*) AS c FROM detection_history{where_clause}", params).fetchone()["c"]
        rows = conn.execute(
            f"SELECT * FROM detection_history{where_clause} ORDER BY time DESC LIMIT ? OFFSET ?",
            params + [per_page, offset],
        ).fetchall()
    finally:
        conn.close()

    return {"success": True, "records": [dict(row) for row in rows], "total": total, "page": page}


@router.post("/history")
async def add_history(request: Request):
    data = await _read_json(request)

    record_id = str(uuid.uuid4())[:8]
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    payload = (
        record_id,
        _to_float(data.get("lat", 0), 0.0),
        _to_float(data.get("lon", 0), 0.0),
        str(data.get("time", now)),
        str(data.get("gas_type", "CH4")),
        str(data.get("result", "normal")),
        _to_float(data.get("concentration", 0), 0.0),
        str(data.get("unit", "ppb")),
        str(data.get("rgb_image", "")),
        str(data.get("mask_image", "")),
        str(data.get("operator", "")),
        str(data.get("remark", "")),
    )

    conn = get_db()
    try:
        conn.execute(
            """
            INSERT INTO detection_history
            (id, lat, lon, time, gas_type, result, concentration, unit, rgb_image, mask_image, operator, remark)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            payload,
        )
        conn.commit()
        row = conn.execute("SELECT * FROM detection_history WHERE id=?", (record_id,)).fetchone()
    finally:
        conn.close()

    return {"success": True, "record": dict(row)}


@router.delete("/history/{record_id}")
async def delete_history(record_id: str, operator: str = Query("", description="Operator username")):
    operator = operator.strip()
    if not operator:
        return _error("Operator is required.", 400)

    conn = get_db()
    try:
        row = conn.execute("SELECT * FROM detection_history WHERE id=?", (record_id,)).fetchone()
        if not row:
            return _error("Record not found.", 404)
        if not _is_admin(operator) and row["operator"] != operator:
            return _error("Permission denied.", 403)

        conn.execute("DELETE FROM detection_history WHERE id=?", (record_id,))
        conn.commit()
    finally:
        conn.close()

    return {"success": True, "message": "Record deleted."}


@router.get("/stats")
async def system_stats():
    conn = get_db()
    try:
        stats = {
            "total_users": conn.execute("SELECT COUNT(*) AS c FROM users").fetchone()["c"],
            "total_devices": conn.execute("SELECT COUNT(*) AS c FROM devices").fetchone()["c"],
            "online_devices": conn.execute("SELECT COUNT(*) AS c FROM devices WHERE status='online'").fetchone()["c"],
            "total_alarms": conn.execute("SELECT COUNT(*) AS c FROM alarms").fetchone()["c"],
            "pending_alarms": conn.execute("SELECT COUNT(*) AS c FROM alarms WHERE status='pending'").fetchone()["c"],
            "resolved_alarms": conn.execute("SELECT COUNT(*) AS c FROM alarms WHERE status='resolved'").fetchone()["c"],
            "total_detections": conn.execute("SELECT COUNT(*) AS c FROM detection_history").fetchone()["c"],
            "anomaly_detections": conn.execute(
                "SELECT COUNT(*) AS c FROM detection_history WHERE result='anomaly'"
            ).fetchone()["c"],
        }
    finally:
        conn.close()

    return {"success": True, "stats": stats}


@router.get("/export/alarms")
async def export_alarms():
    conn = get_db()
    try:
        rows = conn.execute("SELECT * FROM alarms ORDER BY time DESC").fetchall()
    finally:
        conn.close()
    return await _export_excel(
        data=[dict(row) for row in rows],
        headers=["ID", "Station", "Value", "Threshold", "Unit", "Time", "Status", "Handler", "Handled Time", "Remark"],
        keys=["id", "station", "value", "threshold", "unit", "time", "status", "handler", "handle_time", "remark"],
        filename="alarm_report.xlsx",
        sheet_name="alarms",
    )


@router.get("/export/history")
async def export_history():
    conn = get_db()
    try:
        rows = conn.execute("SELECT * FROM detection_history ORDER BY time DESC").fetchall()
    finally:
        conn.close()
    return await _export_excel(
        data=[dict(row) for row in rows],
        headers=["ID", "Latitude", "Longitude", "Time", "Gas", "Result", "Concentration", "Unit", "Operator", "Remark"],
        keys=["id", "lat", "lon", "time", "gas_type", "result", "concentration", "unit", "operator", "remark"],
        filename="detection_history.xlsx",
        sheet_name="history",
    )


@router.get("/export/weekly")
async def export_weekly():
    end = datetime.now()
    start = end - timedelta(days=7)
    start_str = start.strftime("%Y-%m-%d")

    conn = get_db()
    try:
        rows = conn.execute("SELECT * FROM alarms WHERE time>=? ORDER BY time DESC", (start_str,)).fetchall()
    finally:
        conn.close()

    return await _export_excel(
        data=[dict(row) for row in rows],
        headers=["Station", "Value", "Threshold", "Time", "Status", "Handler"],
        keys=["station", "value", "threshold", "time", "status", "handler"],
        filename=f"weekly_report_{end.strftime('%Y%m%d')}.xlsx",
        sheet_name="weekly",
    )


async def _export_excel(data: list[dict[str, Any]], headers: list[str], keys: list[str], filename: str, sheet_name: str):
    os.makedirs(DATA_DIR, exist_ok=True)
    export_path = os.path.join(DATA_DIR, filename)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name

        header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font

        for row_idx, row in enumerate(data, start=2):
            for col_idx, key in enumerate(keys, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=row.get(key, ""))

        for i in range(len(headers)):
            sheet.column_dimensions[chr(65 + i)].width = 16

        workbook.save(export_path)
        return FileResponse(
            export_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=filename,
        )
    except ImportError:
        csv_path = export_path.replace(".xlsx", ".csv")
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for row in data:
            writer.writerow([row.get(key, "") for key in keys])
        with open(csv_path, "wb") as file_obj:
            file_obj.write(output.getvalue().encode("utf-8-sig"))

        return FileResponse(csv_path, media_type="text/csv", filename=os.path.basename(csv_path))
